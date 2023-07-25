import sys
from dronekit import connect, VehicleMode
import argparse
import time
from openpyxl import Workbook,load_workbook


def connectMyPlane():
    parser = argparse.ArgumentParser(description='commands')
    parser.add_argument('--connect', default='tcp:127.0.0.1:5762')
    args = parser.parse_args()
    connection_string = args.connect
    baud_rate = 57600
    vehicle = connect(connection_string, baud=baud_rate, wait_ready=True)
    return vehicle

def function():
    vehicle = connectMyPlane()
    vehicle.mode = VehicleMode("AUTO")
    vehicle.armed = True
    wb = Workbook()
    ws = wb.active
    ws.delete_rows(1, ws.max_row)
    ws.title = "DATA"
    ws["A1"] = "SANİYE"
    ws["B1"] = "HIZ"
    ws["C1"] = "ALTİTUDE"
    ws["D1"] = "PİTCH"
    ws["E1"] = "ROLL"
    ws["F1"] = "YAW"
    ws["G1"] = "BATARYA SEVİYESİ"
    saniye = 0
    while True:
        saniye = saniye + 1
        airspeed = vehicle.airspeed
        altitude = vehicle.location.global_relative_frame.alt
        batarya_seviye = vehicle.battery.level
        attitude = vehicle.attitude
        ws.append([saniye, airspeed, altitude, attitude.pitch, attitude.roll, attitude.yaw, batarya_seviye])
        wb.save("data.xlsx")
        time.sleep(1)

if __name__ == "__main__":
    function()